#!/usr/bin/env python3
"""
Build src/data/stats.json — the four character ratings (Wisdom, Charisma,
Courage, Discipline), each with their full curve from rating −3 .. +15
in both Base and Competitive Mode (CM) variants.

Source of truth:
  - reference/XML/Infos/rating.xml          → id, color, signature yield,
                                              gain/loss flavor lines, icon
  - reference/XML/Infos/text-infos.xml      → Name + Help text per rating
  - Old World Reference Spreadsheet.xlsx    → the per-level curve numbers,
                                              one block per stat × mode

The curve numbers (per-level yields, percent modifiers, crit chance, etc.)
come from a game algorithm that isn't expressed in XML — Soren's hand-
computed table in the legacy spreadsheet is the canonical reference, so
we seed from the xlsx and keep XML-derived metadata wrapped around it.

Output shape:
{
  "wisdom":     { id, slug, name, yield, help, gainText[], loseText[],
                  columns[ {key, label, header} ],
                  base: [ {rating, cells: [v, …]} ],
                  cm:   [ {rating, cells: [v, …]} ] },
  "charisma":   {...},  "courage": {...},  "discipline": {...}
}
"""
from __future__ import annotations

import json
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XML_DIR = ROOT / "reference" / "XML" / "Infos"
XLSX = ROOT / "Old World Reference Spreadsheet.xlsx"
OUT = ROOT / "src" / "data" / "stats.json"

# Map each stat to its primary yield (drives page tint + col defaults).
STAT_YIELD = {
    "wisdom":     "science",
    "charisma":   "civics",
    "courage":    "training",
    "discipline": "money",
}

# Five roles (Leader / Governor / General / Ambassador / Chancellor) in
# spreadsheet column order C..G. Per stat, each role pays out a different
# kind of effect (a yield, a %, a combat bonus, an opinion modifier, …),
# so the column header per stat is data-driven.
ROLE_KEYS = ["leader", "governor", "general", "ambassador", "chancellor"]
ROLE_LABELS = ["Leader", "Governor", "General", "Ambassador", "Chancellor"]

# Spreadsheet sheet names per stat + mode.
SHEET_NAMES = {
    ("wisdom",     "base"): "🟣 Wisdom Base",
    ("wisdom",     "cm"):   "🟣 Wisdom CM",
    ("charisma",   "base"): "🧡 Charisma Base",
    ("charisma",   "cm"):   "🧡 Charisma CM",
    ("courage",    "base"): "🔺 Courage Base",
    ("courage",    "cm"):   "🔺 Courage CM",
    ("discipline", "base"): "⚡ Discipline Base",
    ("discipline", "cm"):   "⚡ Discipline CM",
}

# Per-column "what this means" — pulled directly from the C4..G4 header row
# of each sheet. We re-state them in code (not just read from cells) so the
# render layer can attach yield colors and units deterministically.
#
# Each entry: (heading, unit, classify) where:
#   - heading: short label rendered under the role name
#   - unit:    'flat' (e.g. "+3 Sci"), 'pct' (e.g. "+20%"),
#              'raw'  (the value is already an enriched string)
#   - classify: a yield key to colorize this cell, or None to let
#               classifyYield decide on the text
RATING_COLUMNS = {
    "wisdom": [
        # Leader gets per-court Science
        ("Science",    "flat", "science"),
        ("Science %",  "pct",  "science"),
        ("Crit %",     "pct",  "training"),  # Generals: critical hit chance
        ("Culture",    "flat", "culture"),
        ("Growth",     "flat", "growth"),
    ],
    "charisma": [
        ("Civics",          "flat", "civics"),
        ("Civics %",        "pct",  "civics"),
        ("Defense %",       "pct",  "training"),  # Generals: city defense / attack penalty
        ("Foreign Opinion", "flat", "influence"),
        ("Civics (Family)", "flat", "civics"),    # Chancellor: per-family civics (MP note)
    ],
    "courage": [
        ("Training",     "flat", "training"),
        ("Training %",   "pct",  "training"),
        ("Attack %",     "pct",  "training"),
        ("Tribe Opinion","flat", "wrath"),
        ("Training",     "flat", "training"),
    ],
    "discipline": [
        ("Money",            "flat", "money"),
        ("Money %",          "pct",  "money"),
        ("XP / Year",        "pct",  "training"),    # Generals: XP/year for units
        ("Religion Opinion", "flat", "divine_favor"),
        ("Money",            "flat", "money"),
    ],
}


# ────────────────────────────────────────────────────────────────────────────
# XML readers
# ────────────────────────────────────────────────────────────────────────────

def parse(name: str) -> ET.Element:
    return ET.parse(XML_DIR / name).getroot()


def load_text(*filenames: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for fn in filenames:
        p = XML_DIR / fn
        if not p.exists():
            continue
        for e in ET.parse(p).getroot().findall("Entry"):
            k = e.findtext("zType") or ""
            en = (e.findtext("en-US") or "").split("~")[0].strip()
            if k:
                out[k] = en
    return out


def _strip_link_markup(s: str) -> str:
    """Strip {lowercase:link(TOKEN,N)} and link(TOKEN,N) wrappers, keep the
    human-readable token. e.g. 'link(YIELD_SCIENCE,1)' → 'Science'."""
    import re

    def repl(m: "re.Match[str]") -> str:
        token = m.group(1)
        return token.replace("YIELD_", "").replace("CONCEPT_", "").replace("_", " ").title()

    s = re.sub(r"\{lowercase:link\(([A-Z_]+),\d+\)\}", repl, s)
    s = re.sub(r"link\(([A-Z_]+),\d+\)", repl, s)
    return s


def read_ratings() -> dict[str, dict]:
    """Return canonical rating metadata keyed by lowercased stat name."""
    text = load_text("text-infos.xml", "text-misc.xml", "text-new.xml")
    out: dict[str, dict] = {}
    for e in parse("rating.xml").findall("Entry"):
        zt = e.findtext("zType") or ""
        if not zt.startswith("RATING_"):
            continue
        slug = zt.replace("RATING_", "").lower()
        name = text.get(e.findtext("Name") or "", slug.title())
        help_raw = text.get(e.findtext("Help") or "", "")
        help_clean = _strip_link_markup(help_raw)
        # Signature yield: from aiYieldCourtRate (Leader-role yield)
        sig_yield = ""
        for pair in e.findall("aiYieldCourtRate/Pair"):
            sig_yield = (pair.findtext("zIndex") or "").replace("YIELD_", "").lower()
            break
        # Per-rating-point modifiers (the raw "what does +1 rating do" values
        # straight from XML, complementary to the spreadsheet curve).
        per_point: list[str] = []
        crit = e.findtext("iCriticalChance")
        if crit and crit != "0":
            per_point.append(f"+{crit}% Critical Hit per point (General)")
        att = e.findtext("iAttackModifier")
        if att and att != "0":
            per_point.append(f"+{att}% Attack per point (General)")
        defm = e.findtext("iDefenseModifier")
        if defm and defm != "0":
            per_point.append(f"+{defm}% Defense per point (General)")
        xp = e.findtext("iUnitXP")
        if xp and xp != "0":
            per_point.append(f"+{xp} Unit XP/Year per point (General)")
        out[slug] = {
            "id": zt,
            "slug": slug,
            "name": name,
            "yield": sig_yield,
            "help": help_clean,
            "perPoint": per_point,
            "color": e.findtext("zColor") or "",
            "icon": e.findtext("zIconName") or "",
        }
    return out


# ────────────────────────────────────────────────────────────────────────────
# Spreadsheet reader
# ────────────────────────────────────────────────────────────────────────────

def _coerce(v):
    """Spreadsheet cells can be: None, int, float, or str. Keep strings as-is
    (e.g. '-6% attack'); keep numbers as numbers."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        # Round to 2 decimals to avoid 0.1000000001 noise.
        if isinstance(v, float):
            v = round(v, 2)
            if v == int(v):
                v = int(v)
        return v
    return str(v).strip()


def read_curve(slug: str, mode: str) -> list[dict]:
    """Read rows 5..23 (rating −3 .. +15) of a given sheet. Returns a list
    of {rating, cells: [c, d, e, f, g]} dicts."""
    from openpyxl import load_workbook  # local import — only needed at build time

    wb = load_workbook(XLSX, data_only=True)
    sheet_name = SHEET_NAMES[(slug, mode)]
    ws = wb[sheet_name]
    rows: list[dict] = []
    for r in range(5, 24):  # 19 rows: ratings −3 .. +15
        rating = _coerce(ws.cell(r, 1).value)
        if rating is None:
            continue
        cells = [_coerce(ws.cell(r, c).value) for c in (3, 4, 5, 6, 7)]
        rows.append({"rating": rating, "cells": cells})
    return rows


# ────────────────────────────────────────────────────────────────────────────
# Build
# ────────────────────────────────────────────────────────────────────────────

def build_stat(slug: str, rating_meta: dict) -> dict:
    cols_spec = RATING_COLUMNS[slug]
    columns = [
        {
            "role":     ROLE_KEYS[i],
            "label":    ROLE_LABELS[i],
            "header":   heading,
            "unit":     unit,
            "yieldKey": classify,
        }
        for i, (heading, unit, classify) in enumerate(cols_spec)
    ]
    return {
        **rating_meta,
        "yieldKey": STAT_YIELD[slug],
        "columns":  columns,
        "base":     read_curve(slug, "base"),
        "cm":       read_curve(slug, "cm"),
    }


def main() -> int:
    ratings = read_ratings()
    out: dict[str, dict] = {}
    for slug in ("wisdom", "charisma", "courage", "discipline"):
        meta = ratings.get(slug)
        if meta is None:
            print(f"!! rating.xml missing RATING_{slug.upper()}", file=sys.stderr)
            continue
        out[slug] = build_stat(slug, meta)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, indent=2, sort_keys=True, ensure_ascii=False) + "\n")
    rows_each = {k: len(v["base"]) for k, v in out.items()}
    print(f"✓ wrote {OUT.relative_to(ROOT)} — 4 ratings, rows per stat: {rows_each}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
